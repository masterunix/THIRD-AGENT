import os
import time
import json
import uuid
import threading
import openpyxl
import requests as http_requests
from datetime import datetime
from flask import Blueprint, request, jsonify
from typing import TypedDict, List, Dict, Any

from langgraph.graph import StateGraph, END

level3_bp = Blueprint('level3', __name__)

VECTORSTORE = None
LLM = None
GEMINI_API_KEY = os.getenv('GEMINI_API_KEY', '')
PAI_API_KEY = os.getenv('PAI_API_KEY', '')

def setup_level3(vectorstore, llm):
    global VECTORSTORE, LLM
    VECTORSTORE = vectorstore
    LLM = llm

# State Definition
class GraphState(TypedDict):
    shipment: dict
    events: list
    alternatives: list
    original_alternatives: list
    current_node: str
    severity: str
    delay_impact: float
    decision: dict
    retry_count: int
    relaxed_criteria: bool
    processing_duration: float
    provider: str

# Global variables for workflow tracking
active_workflows = {}
audit_logs = {}
custom_dataset_path = None  # For custom uploaded files

def call_llm(prompt: str, system_prompt: str = "", provider: str = "azure") -> str:
    """Call LLM with multi-provider support - optimized for speed"""
    if provider == "azure" and LLM:
        try:
            from langchain_core.messages import SystemMessage, HumanMessage
            messages = []
            if system_prompt:
                messages.append(SystemMessage(content=system_prompt))
            messages.append(HumanMessage(content=prompt))
            # Set shorter timeout for faster responses
            response = LLM.invoke(messages, config={"timeout": 10, "max_tokens": 50})
            return response.content
        except Exception as e:
            print(f"Azure LLM error: {e}")
            # Fallback to Gemini
            result = call_llm_gemini(prompt, system_prompt)
            if result:
                return result
            # Fallback to PAI
            return call_llm_pai(prompt, system_prompt)
    elif provider == "gemini":
        result = call_llm_gemini(prompt, system_prompt)
        if result:
            return result
        # Fallback to Azure
        if LLM:
            try:
                from langchain_core.messages import SystemMessage, HumanMessage
                messages = []
                if system_prompt:
                    messages.append(SystemMessage(content=system_prompt))
                messages.append(HumanMessage(content=prompt))
                response = LLM.invoke(messages)
                return response.content
            except:
                pass
        # Fallback to PAI
        return call_llm_pai(prompt, system_prompt)
    elif provider == "pai":
        result = call_llm_pai(prompt, system_prompt)
        if result:
            return result
        # Fallback to Azure
        if LLM:
            try:
                from langchain_core.messages import SystemMessage, HumanMessage
                messages = []
                if system_prompt:
                    messages.append(SystemMessage(content=system_prompt))
                messages.append(HumanMessage(content=prompt))
                response = LLM.invoke(messages)
                return response.content
            except:
                pass
        # Fallback to Gemini
        return call_llm_gemini(prompt, system_prompt)
    
    return "LLM call failed - all providers unavailable"

def call_llm_gemini(prompt: str, system_prompt: str = "") -> str:
    """Call Gemini API with optimized settings"""
    if not GEMINI_API_KEY:
        return ""
    try:
        url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={GEMINI_API_KEY}"
        full_prompt = f"{system_prompt}\n\n{prompt}" if system_prompt else prompt
        payload = {
            "contents": [
                {"role": "user", "parts": [{"text": full_prompt}]}
            ],
            "generationConfig": {
                "temperature": 0.3,  # Lower for faster, more focused responses
                "maxOutputTokens": 50,  # Limit output length
                "thinkingConfig": {"thinkingBudget": 0}
            }
        }
        resp = http_requests.post(url, json=payload, timeout=10)  # Reduced timeout
        if resp.status_code == 200:
            data = resp.json()
            text = data.get('candidates', [{}])[0].get('content', {}).get('parts', [{}])[0].get('text', '')
            return text
        print(f"Gemini API returned {resp.status_code}")
        return ""
    except Exception as e:
        print(f"Gemini API error: {e}")
        return ""

def call_llm_pai(prompt: str, system_prompt: str = "") -> str:
    """Call PAI API with optimized settings"""
    if not PAI_API_KEY:
        return ""
    try:
        headers = {
            "Authorization": f"Bearer {PAI_API_KEY}",
            "Content-Type": "application/json"
        }
        payload = {
            "model": "gemma4:26b",
            "messages": [
                {"role": "system", "content": system_prompt} if system_prompt else None,
                {"role": "user", "content": prompt}
            ],
            "max_tokens": 50,  # Limit response length
            "temperature": 0.3  # Lower for faster responses
        }
        # Remove None values
        payload["messages"] = [m for m in payload["messages"] if m is not None]
        
        resp = http_requests.post("https://pai-api.thepsi.com/api/v4/chat", json=payload, headers=headers, timeout=30)  # Increased timeout
        if resp.status_code == 200:
            lines = resp.text.strip().split('\n')
            final_text = ""
            for line in lines:
                if not line:
                    continue
                try:
                    data = json.loads(line)
                    if data.get('type') == 'text':
                        final_text = data.get('content', '')
                except:
                    pass
            return final_text
        print(f"PAI API returned {resp.status_code}")
        return ""
    except Exception as e:
        print(f"PAI API error: {e}")
        return ""

def load_excel_dataset(filepath="LEVEL3FILES/shipments_dataset.xlsx"):
    if not os.path.exists(filepath):
        raise ValueError(f"Dataset file not found: {filepath}")
        
    import re
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = {}
    
    required_sheets = ['Active_Shipments', 'Disruption_Events', 'Alternative_Routes', 'SLA_Reference']
    
    for sheet_name in wb.sheetnames:
        if sheet_name not in required_sheets and sheet_name != 'Scoring_Guide':
            continue
            
        ws = wb[sheet_name]
        records = []
        
        # The first row is title, second row is header
        # Let's find the header row by looking for the row with actual headers
        # Assuming row 2 is header (1-indexed, so row=2 in openpyxl)
        
        headers = [cell.value for cell in ws[2]]
        
        for row in ws.iter_rows(min_row=3, values_only=True):
            # Check if row is completely empty
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
                
            record = {}
            for key, value in zip(headers, row):
                if key is None:
                    continue
                    
                if isinstance(value, datetime):
                    record[key] = str(value)
                else:
                    record[key] = value
                    
            if sheet_name == 'Active_Shipments' and 'Shipment ID' in record:
                shipment_id = str(record.get('Shipment ID', ''))
                # Filter out legend/colour-key rows (not real shipments)
                if not re.match(r'^[A-Za-z].*\d', shipment_id):
                    continue
                    
            records.append(record)
            
        sheets[sheet_name] = records
        
    for sheet in required_sheets:
        if sheet not in sheets:
            raise ValueError(f"Required sheet missing: {sheet}")
            
    if 'Scoring_Guide' not in sheets:
        sheets['Scoring_Guide'] = []
        
    return sheets

def extract_field(record, *keys):
    """Robustly extract a field from a record, trying multiple key aliases (case-insensitive)."""
    lower = {str(k).lower().strip(): v for k, v in record.items()}
    for key in keys:
        val = lower.get(key.lower().strip())
        if val is not None and str(val).strip() not in ('', 'nan', 'none'):
            return val
    return None

def get_shipment_id(record, fallback_idx=None):
    """Extract shipment ID from a record using all known column aliases."""
    val = extract_field(record,
        'shipment id', 'shipment_id', 'shipmentid', 'id',
        'shipment no', 'shipment number', 'shipment_no'
    )
    if val is not None:
        return str(val).strip()
    return f"UNKNOWN_{fallback_idx}" if fallback_idx is not None else None

def prioritize_shipments(shipments):
    tier_order = {"Platinum": 1, "Gold": 2, "Silver": 3, "Bronze": 4}
    status_order = {"grounded": 1, "delayed": 2, "in_transit": 3}
    def get_tier(x):
        t = extract_field(x, 'sla tier', 'tier', 'customer tier', 'priority')
        return tier_order.get(str(t).strip() if t else '', 5)
    def get_status(x):
        s = extract_field(x, 'current status', 'status', 'shipment status', 'shipment_status')
        return status_order.get(str(s).strip().lower() if s else '', 5)
    def get_value(x):
        v = extract_field(x, 'cargo value (usd)', 'cargo_value_usd', 'cargo value', 'value')
        try:
            return -float(v or 0)
        except (ValueError, TypeError):
            return 0
    return sorted(shipments, key=lambda x: (get_tier(x), get_status(x), get_value(x)))

# --- Nodes ---
def assess_impact(state: GraphState):
    events = state['events']
    shipment = state['shipment']
    
    # Short-circuit: ON_TRACK shipments with no events need no action
    current_status = str(extract_field(shipment, 'current status', 'status') or '').upper()
    if current_status == 'ON_TRACK' and not events:
        return {"current_node": "assess_impact", "delay_impact": 0.0, "severity": "NONE"}
    
    # Support both original column names and normalised aliases
    def get_delay(e):
        v = extract_field(e, 'delay (hrs)', 'delay_hours', 'delay hours', 'delay(hrs)', 'additional_delay_hours')
        try:
            return float(v or 0)
        except (ValueError, TypeError):
            return 0.0
    total_delay = sum(get_delay(e) for e in events)
    
    # Support both normalised and original tolerance column names
    tol_val = extract_field(shipment, 'sla tolerance (hrs)', 'sla_tolerance_hours', 'sla tolerance', 'tolerance_hours')
    try:
        tolerance = float(tol_val or 24)
    except (ValueError, TypeError):
        tolerance = 24.0
    
    if total_delay > tolerance:
        severity = "CRITICAL"
    elif total_delay >= 0.5 * tolerance:
        severity = "HIGH"
    elif total_delay >= 0.25 * tolerance:
        severity = "MEDIUM"
    else:
        severity = "LOW"
    
    return {"current_node": "assess_impact", "delay_impact": total_delay, "severity": severity}

def score_alternatives(state: GraphState):
    alts = state['alternatives']
    shipment = state['shipment']
    cargo_type = str(extract_field(shipment, 'cargo type', 'cargo_type') or '').lower()
    tier = str(extract_field(shipment, 'sla tier', 'tier', 'customer tier') or '').strip()
    
    relaxed = state.get('relaxed_criteria', False)
    
    def safe_float(record, *keys, default=50.0):
        v = extract_field(record, *keys)
        try:
            return float(v) if v is not None else default
        except (ValueError, TypeError):
            return default
    
    needs_cold_chain = "pharma" in cargo_type or "medical" in cargo_type or "vaccine" in cargo_type or "cold chain" in cargo_type
    
    scored = []
    for alt in alts:
        avail   = safe_float(alt, 'availability score', 'availability_score', 'availability', default=50.0)
        cost    = safe_float(alt, 'cost score', 'cost_score', 'cost', default=50.0)
        transit = safe_float(alt, 'transit score', 'transit_score', 'transit time (days)', 'transit days', default=50.0)
        rel     = safe_float(alt, 'reliability score', 'reliability_score', 'reliability', default=50.0)
        surcharge = safe_float(alt, 'surcharge (%)', 'surcharge_percent', 'surcharge(%)', 'surcharge (usd)', default=0.0)
        
        # Cold-chain constraint: check multiple possible column representations
        if needs_cold_chain:
            # Check dedicated boolean column
            cc_bool = extract_field(alt, 'cold chain certified', 'cold_chain_certified', 'cold chain', 'cold-chain')
            # Check availability text for cold-chain keywords
            avail_text = str(extract_field(alt, 'availability', 'notes', 'availability score') or '').lower()
            # Check alt route description for cold-chain keywords
            alt_route_text = str(extract_field(alt, 'alt route', 'alternative route', 'route') or '').lower()
            # Check alt carrier name for known cold-chain carriers
            carrier_name = str(extract_field(alt, 'alt carrier', 'alternative carrier', 'carrier') or '').lower()
            cold_chain_carriers = ['singapore airlines', 'air india cargo', 'lufthansa cargo', 'emirates skycargo', 'cathay pacific cargo', 'iac', 'air france cargo', 'brussels airlines', 'british airways']
            carrier_has_cc = any(c in carrier_name for c in cold_chain_carriers)
            cc_in_text = 'cold-chain' in avail_text or 'cold chain' in avail_text or 'cold-chain' in alt_route_text or 'cold chain' in alt_route_text
            cc_flag = str(cc_bool or '').strip().lower() in ('true', 'yes', '1')
            if not (cc_flag or cc_in_text or carrier_has_cc):
                continue
        
        # Surcharge constraint by tier - calculate as percentage of cargo value
        cargo_value = safe_float(shipment, 'cargo value (usd)', 'cargo_value_usd', 'cargo value', 'value', default=100000.0)
        surcharge_percent = (surcharge / cargo_value) * 100 if cargo_value > 0 else 0
        
        if not relaxed:
            if tier == 'Platinum' and surcharge_percent > 20:
                continue
        else:
            if tier == 'Platinum' and surcharge_percent > 30:
                continue
        
        comp = (avail * 0.3) + (cost * 0.3) + (transit * 0.2) + (rel * 0.2)
        scored.append({"alt": alt, "score": comp})
    
    scored.sort(key=lambda x: x['score'], reverse=True)
    return {"current_node": "score_alternatives", "alternatives": [s['alt'] for s in scored]}

def decide(state: GraphState):
    alts = state['alternatives']
    severity = state.get('severity', 'LOW')
    
    # ON_TRACK shipments with no delay: no action needed
    if severity == 'NONE' and not alts:
        return {"current_node": "decide", "decision": {
            "outcome": "NO_ACTION_NEEDED",
            "reasoning": "Shipment is ON_TRACK with no disruption events. Monitor only."
        }}
    
    if not alts:
        return {"current_node": "decide", "decision": {"outcome": "RELAXED_CRITERIA_NEEDED" if state['retry_count'] == 0 else "ESCALATED"}}
    
    best_alt = alts[0]
    carrier = extract_field(best_alt, 'alt carrier', 'alternative carrier', 'carrier', 'recommended carrier') or 'alternative carrier'
    outcome = "REROUTED_RELAXED" if state.get('relaxed_criteria') else "REROUTED"
    return {"current_node": "decide", "decision": {
        "outcome": outcome,
        "selected_route": best_alt,
        "reasoning": f"Selected highest scoring route with carrier {carrier}"
    }}

def retry_with_relaxed_criteria(state: GraphState):
    return {
        "current_node": "retry_with_relaxed_criteria",
        "retry_count": state.get('retry_count', 0) + 1,
        "relaxed_criteria": True,
        "alternatives": state.get('original_alternatives', [])
    }

def escalate(state: GraphState):
    shipment = state['shipment']
    cargo_type = str(extract_field(shipment, 'cargo type', 'cargo_type') or '').lower()
    delay = state['delay_impact']
    notes = str(extract_field(shipment, 'notes / context', 'notes', 'context', 'notes_context') or '').lower()
    
    reason = "NO_VIABLE_ROUTE"
    if ("pharma" in cargo_type or "medical" in cargo_type) and delay > 2:
        reason = "CRITICAL_CARGO"
    elif "regulatory hold" in notes or "customs hold" in notes:
        reason = "REGULATORY_HOLD"
    
    return {"current_node": "escalate", "decision": {
        "outcome": "ESCALATED",
        "reason": reason,
        "reasoning": f"Escalated due to {reason}"
    }}

def decide_router(state: GraphState):
    outcome = state['decision'].get('outcome')
    if outcome in ('NO_ACTION_NEEDED',):
        return END
    if outcome == "RELAXED_CRITERIA_NEEDED":
        if state['retry_count'] >= 1:
            return "escalate"
        return "retry_with_relaxed_criteria"
    if outcome == "ESCALATED":
        return "escalate"
    return END

# Build Graph
workflow = StateGraph(GraphState)
workflow.add_node("assess_impact", assess_impact)
workflow.add_node("score_alternatives", score_alternatives)
workflow.add_node("decide", decide)
workflow.add_node("retry_with_relaxed_criteria", retry_with_relaxed_criteria)
workflow.add_node("escalate", escalate)

workflow.set_entry_point("assess_impact")
workflow.add_edge("assess_impact", "score_alternatives")
workflow.add_edge("score_alternatives", "decide")
workflow.add_conditional_edges("decide", decide_router)
workflow.add_edge("retry_with_relaxed_criteria", "score_alternatives")
workflow.add_edge("escalate", END)

app_graph = workflow.compile()

def process_single_shipment(shipment, global_idx, dataset, wf, provider, workflow_id):
    if wf['status'] == 'CANCELLED':
        return False
        
    sh_id = get_shipment_id(shipment, fallback_idx=global_idx)
        
    wf['current_shipment_id'] = sh_id
    wf['processing_status'] = 'PROCESSING'
    
    start_t = time.time()
    
    # Robustly match IDs for events and alternatives
    def match_id(record, target_id):
        rec_id = extract_field(record,
            'shipment id', 'shipment_id', 'shipmentid', 'id',
            'affected shipment', 'affected shipment id', 'affected_shipment',
            'shipment', 'shipment_no', 'shipment number', 'shipment no'
        )
        return rec_id is not None and str(rec_id).strip() == str(target_id).strip()

    alts = [a for a in dataset['Alternative_Routes'] if match_id(a, sh_id)]
    state = {
        "shipment": shipment,
        "events": [e for e in dataset['Disruption_Events'] if match_id(e, sh_id)],
        "alternatives": alts,
        "original_alternatives": alts,
        "current_node": "assess_impact",
        "severity": "LOW",
        "delay_impact": 0.0,
        "decision": {},
        "retry_count": 0,
        "relaxed_criteria": False,
        "processing_duration": 0.0,
        "provider": provider
    }
    
    try:
        final_decision = {}
        for s in app_graph.stream(state, {"recursion_limit": 10}):
            node_name = list(s.keys())[0]
            wf['current_node'] = node_name
            
            node_state = s[node_name]
            if isinstance(node_state, dict) and 'decision' in node_state:
                final_decision = node_state['decision']
            
            audit_logs[workflow_id].append({
                "timestamp": datetime.now().isoformat(),
                "shipment_id": sh_id,
                "node": node_name
            })
        
        dur = time.time() - start_t
        outcome = final_decision.get('outcome', 'PROCESSED')
        
        if outcome in ['ESCALATED', 'REROUTED_RELAXED']:
            llm_prompt = f"{sh_id} ({shipment.get('tier')}): {outcome}. Why?"
            rationale = call_llm(llm_prompt, "Explain in 10 words max.", provider)
            if rationale:
                final_decision['llm_rationale'] = rationale.strip()
            else:
                final_decision['llm_rationale'] = f"{outcome}: {final_decision.get('reason', 'See details')}"
        else:
            if final_decision.get('selected_route'):
                carrier = final_decision['selected_route'].get('carrier', 'alternative')
                final_decision['llm_rationale'] = f"Optimal route via {carrier}"
            else:
                final_decision['llm_rationale'] = "Route optimized successfully"
        
        wf['decisions'][sh_id] = {
            "outcome": outcome,
            "duration": dur,
            "provider": provider,
            "reasoning": final_decision.get('reasoning', ''),
            "llm_rationale": final_decision.get('llm_rationale', ''),
            **final_decision
        }
    except Exception as e:
        print(f"Workflow error for {sh_id}:", e)
        wf['decisions'][sh_id] = {
            "outcome": "ERROR",
            "reason": str(e),
            "duration": time.time() - start_t,
            "provider": provider
        }
    return True

def run_workflow_background(workflow_id, provider='azure'):
    from concurrent.futures import ThreadPoolExecutor
    wf = active_workflows[workflow_id]
    dataset = wf['dataset']
    queue = wf['queue']
    
    audit_logs[workflow_id] = []
    
    BATCH_SIZE = 5
    
    with ThreadPoolExecutor(max_workers=5) as executor:
        for batch_start in range(0, len(queue), BATCH_SIZE):
            batch_end = min(batch_start + BATCH_SIZE, len(queue))
            batch = queue[batch_start:batch_end]
            
            if wf['status'] == 'CANCELLED':
                break
                
            futures = []
            for idx_in_batch, shipment in enumerate(batch):
                global_idx = batch_start + idx_in_batch
                futures.append(executor.submit(process_single_shipment, shipment, global_idx, dataset, wf, provider, workflow_id))
            
            for future in futures:
                future.result()
                
            wf['current_shipment_index'] = batch_end
            time.sleep(0.05)
            
            if batch_end < len(queue):
                time.sleep(0.1)

    wf['status'] = 'COMPLETED'
    wf['processing_status'] = 'IDLE'

# API Endpoints
@level3_bp.route('/upload-dataset', methods=['POST'])
def upload_dataset():
    """Upload a custom Excel dataset"""
    try:
        if 'file' not in request.files:
            return jsonify({"error": "No file provided"}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({"error": "No file selected"}), 400
        
        if not file.filename.endswith('.xlsx'):
            return jsonify({"error": "File must be an Excel file (.xlsx)"}), 400
        
        # Save to temporary location
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        filepath = os.path.join(upload_dir, f"custom_dataset_{int(time.time())}.xlsx")
        file.save(filepath)
        
        # Validate the file
        try:
            sheets = load_excel_dataset(filepath)
            global custom_dataset_path
            custom_dataset_path = filepath
            
            return jsonify({
                "success": True,
                "filepath": filepath,
                "sheets": list(sheets.keys()),
                "shipment_count": len(sheets['Active_Shipments'])
            })
        except Exception as e:
            os.remove(filepath)
            return jsonify({"error": f"Invalid dataset: {str(e)}"}), 400
            
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@level3_bp.route('/reset-dataset', methods=['POST'])
def reset_dataset():
    """Reset to default dataset"""
    global custom_dataset_path
    if custom_dataset_path and os.path.exists(custom_dataset_path):
        try:
            os.remove(custom_dataset_path)
        except:
            pass
    custom_dataset_path = None
    return jsonify({"success": True})

@level3_bp.route('/dataset-info', methods=['GET'])
def dataset_info():
    """Get information about the current dataset"""
    try:
        filepath = custom_dataset_path if custom_dataset_path else "LEVEL3FILES/shipments_dataset.xlsx"
        sheets = load_excel_dataset(filepath)
        
        return jsonify({
            "filepath": filepath,
            "is_custom": custom_dataset_path is not None,
            "sheets": list(sheets.keys()),
            "shipment_count": len(sheets['Active_Shipments']),
            "disruption_count": len(sheets['Disruption_Events']),
            "alternative_count": len(sheets['Alternative_Routes'])
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@level3_bp.route('/start-workflow', methods=['POST'])
def start_workflow():
    try:
        # Import active_provider from main backend module dynamically
        import backend
        current_provider = backend.active_provider
        
        filepath = custom_dataset_path if custom_dataset_path else "LEVEL3FILES/shipments_dataset.xlsx"
        dataset = load_excel_dataset(filepath)
        queue = prioritize_shipments(dataset['Active_Shipments'])
        
        workflow_id = str(uuid.uuid4())
        active_workflows[workflow_id] = {
            "id": workflow_id,
            "dataset": dataset,
            "queue": queue,
            "status": "RUNNING",
            "processing_status": "STARTING",
            "current_shipment_index": 0,
            "current_shipment_id": None,
            "current_node": None,
            "total_shipments": len(queue),
            "decisions": {},
            "filepath": filepath,
            "provider": current_provider
        }
        
        t = threading.Thread(target=run_workflow_background, args=(workflow_id, current_provider))
        t.start()
        
        # Clean queue for JSON response - ensure no NaN values
        clean_queue = []
        for idx, item in enumerate(queue):
            clean_item = {}
            for key, value in item.items():
                if value is None or (isinstance(value, float) and value != value):
                    clean_item[key] = None
                else:
                    clean_item[key] = value
                    
            # Normalize keys for the frontend using the extract_field helper
            sh_id = get_shipment_id(item, fallback_idx=idx)
            clean_item['shipment_id'] = sh_id
            
            cust = extract_field(item, 'customer', 'customer name', 'client')
            clean_item['customer'] = cust if cust else 'Unknown'
            
            ctype = extract_field(item, 'cargo type', 'cargo_type', 'cargo')
            clean_item['cargo_type'] = ctype if ctype else 'General'
            
            tier = extract_field(item, 'sla tier', 'tier', 'customer tier', 'priority')
            clean_item['tier'] = tier if tier else 'Standard'
            
            status = extract_field(item, 'current status', 'status', 'shipment status', 'shipment_status')
            clean_item['status'] = status if status else 'unknown'
                
            clean_queue.append(clean_item)
        
        return jsonify({
            "workflow_id": workflow_id,
            "status": "RUNNING",
            "total_shipments": len(clean_queue),
            "queue": clean_queue,
            "filepath": filepath,
            "provider": current_provider
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500

@level3_bp.route('/workflow-status/<workflow_id>', methods=['GET'])
def workflow_status(workflow_id):
    if workflow_id not in active_workflows:
        return jsonify({"error": "Workflow not found"}), 404
    
    wf = active_workflows[workflow_id]
    return jsonify({
        "workflow_id": workflow_id,
        "status": wf['status'],
        "processing_status": wf['processing_status'],
        "current_shipment_index": wf['current_shipment_index'],
        "total_shipments": wf['total_shipments'],
        "current_node": wf['current_node'],
        "current_shipment_id": wf['current_shipment_id']
    })

@level3_bp.route('/shipment-decision/<workflow_id>/<shipment_id>', methods=['GET'])
def shipment_decision(workflow_id, shipment_id):
    if workflow_id not in active_workflows:
        return jsonify({"error": "Workflow not found"}), 404
    
    wf = active_workflows[workflow_id]
    if shipment_id not in wf['decisions']:
        return jsonify({"error": "Decision not found"}), 404
        
    return jsonify({
        "shipment_id": shipment_id,
        "decision": wf['decisions'][shipment_id]
    })

@level3_bp.route('/workflow-cancel/<workflow_id>', methods=['POST'])
def cancel_workflow(workflow_id):
    if workflow_id not in active_workflows:
        return jsonify({"error": "Workflow not found"}), 404
    
    active_workflows[workflow_id]['status'] = 'CANCELLED'
    return jsonify({"success": True})

@level3_bp.route('/audit-log/<workflow_id>', methods=['GET'])
def get_audit_log(workflow_id):
    if workflow_id not in audit_logs:
        return jsonify({"error": "Logs not found"}), 404
    return jsonify({"audit_log": audit_logs[workflow_id]})
